import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ===== متغيرات عالمية =====
current_df = None
current_file = None

# ===== فتح ملف Excel =====
def open_excel():
    global current_df, current_file
    path = filedialog.askopenfilename(filetypes=[("Excel files", ".xlsx;.xls")])
    if not path:
        return
    try:
        df = pd.read_excel(path, engine="openpyxl")
    except Exception as e:
        messagebox.showerror("خطأ في فتح الملف", str(e))
        return
    current_df = df
    current_file = path
    show_dataframe(df)
    status.config(text=f"تم فتح الملف: {os.path.basename(current_file)} | {len(current_df)} صفوف")
# ===== عرض البيانات في Treeview =====
def show_dataframe(df):
    for r in tree.get_children():
        tree.delete(r)
    tree["columns"] = list(df.columns)
    tree["show"] = "headings"
    for c in df.columns:
        tree.heading(c, text=c)
        tree.column(c, width=120)
    for row in df.itertuples(index=False):
        tree.insert("", "end", values=tuple(row))
# ===== تنسيق الورقة =====
def format_sheet(ws, header_color="006400", hyperlink_col=None):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal="left")
    if hyperlink_col:
        for row in ws.iter_rows(min_row=2, min_col=hyperlink_col, max_col=hyperlink_col):
            for cell in row:
                if cell.value:
                    cell.font = Font(color="006400", size=12)

def generate_etisalat_report():
    global current_df, current_file
    if current_df is None:
        messagebox.showinfo("مطلوب", "افتح ملف Excel أولاً")
        return

    import os
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # ===== تنظيف الأعمدة =====
    current_df.columns = current_df.columns.str.strip()
    current_df['Originating_Number'] = current_df['Originating_Number'].astype(str)
    current_df['Terminating_Number'] = current_df['Terminating_Number'].astype(str)

    # ===== حساب تكرار الأرقام =====
    numbers = pd.concat([current_df['Originating_Number'], current_df['Terminating_Number']])
    freq = numbers.value_counts().reset_index()
    freq.columns = ['B Number', 'Count']

    df_final = freq.copy()

    # ===== إنشاء dictionary للبيانات لكل رقم =====
    b_data_dict = {}
    for _, row in current_df.iterrows():
        for num_col in ['Originating_Number', 'Terminating_Number']:
            num = str(row[num_col])
            if num not in b_data_dict:
                b_data_dict[num] = {
                    'B Full Name': row['B_Number_Full_Name'],
                    'B Address': row['B_Number_Address'],
                    'B_NUMBER_SITE_ADDRESS': row['B_Number_MU_Site_Address'],
                    'Latitude': row['B_Number_MU_Latitude'],
                    'Longitude': row['B_Number_MU_Longitude']
                }

    # ===== إضافة الأعمدة للـ df_final =====
    df_final['B Full Name'] = df_final['B Number'].map(lambda x: b_data_dict[x]['B Full Name'] if x in b_data_dict else '')
    df_final['B Address'] = df_final['B Number'].map(lambda x: b_data_dict[x]['B Address'] if x in b_data_dict else '')
    df_final['B_NUMBER_SITE_ADDRESS'] = df_final['B Number'].map(lambda x: b_data_dict[x]['B_NUMBER_SITE_ADDRESS'] if x in b_data_dict else '')
    df_final['Latitude'] = df_final['B Number'].map(lambda x: b_data_dict[x]['Latitude'] if x in b_data_dict else '')
    df_final['Longitude'] = df_final['B Number'].map(lambda x: b_data_dict[x]['Longitude'] if x in b_data_dict else '')

    # ===== خانات Map =====
    df_final['Map'] = df_final.apply(
        lambda row: f'=HYPERLINK("https://www.google.com/maps/search/?api=1&query={row["Latitude"]},{row["Longitude"]}","Map")'
        if pd.notna(row['Latitude']) and pd.notna(row['Longitude']) else '', axis=1
    )

    # ===== حساب SMS من Originating_Number =====
    temp_df = current_df.copy()
    temp_df['activity_clean'] = temp_df['Network_Activity_Type_Name'].astype(str).str.strip()
    activity_stats = temp_df.groupby('Originating_Number').agg(
        SMS=('activity_clean', lambda x: (x=="SMS").sum())
    ).reset_index()
    df_final = df_final.merge(activity_stats, left_on='B Number', right_on='Originating_Number', how='left')
    df_final.drop(columns=['Originating_Number'], inplace=True)
    df_final['SMS'] = df_final['SMS'].fillna(0).astype(int)
    df_final['Count'] = df_final['Count'].astype(int)

    # ===== حساب أول وآخر مكالمة لكل B Number =====
    temp_df['Call_Start_Date'] = pd.to_datetime(temp_df['Call_Start_Date'])
    call_records = pd.concat([
        temp_df[['Originating_Number','Call_Start_Date']].rename(columns={'Originating_Number':'B Number'}),
        temp_df[['Terminating_Number','Call_Start_Date']].rename(columns={'Terminating_Number':'B Number'})
    ])
    first_last_call = call_records.groupby('B Number').agg(
        First_Call=('Call_Start_Date','min'),
        Last_Call=('Call_Start_Date','max')
    ).reset_index()
    df_final = df_final.merge(first_last_call, on='B Number', how='left')

    df_final = df_final.sort_values(by='Count', ascending=False)

    # ===== استثناء الصف الأول الأكثر تكرار =====
    if len(df_final) >= 1:
        target_number = freq.iloc[0]['B Number']
        mask = df_final['B Number'] == target_number
        df_final.loc[mask, ['B Full Name','B Address','B_NUMBER_SITE_ADDRESS','Latitude','Longitude','Map','SMS']] = \
            f"{current_df.at[0,'A_Number_Details_First_Name']} {current_df.at[0,'A_Number_Details_Last_Name']}", \
            '28607102800033', current_df.at[0,'MU_Site_Address'], '', '', '', 0

    # ===== معالجة IMEI =====
    def safe_imei(x):
        try:
            return str(int(float(x)))
        except:
            return ''

    imei_df = current_df.copy()
    imei_df['IMEI_Number'] = imei_df['IMEI_Number'].apply(safe_imei)
    imei_summary = imei_df.groupby('IMEI_Number').agg(
        Count=('IMEI_Number','count'),
        First_Use_Date=('Call_Start_Date','min'),
        Last_Use_Date=('Call_Start_Date','max'),
        First_Use_Address=('Site_Address','first'),
        Last_Use_Address=('Site_Address','last')
    ).reset_index()
    imei_summary.rename(columns={'IMEI_Number':'IMEI'}, inplace=True)
    imei_summary['Device Info'] = imei_summary['IMEI'].apply(
        lambda x: f'=HYPERLINK("https://www.imei.info/calc/?imei={x}","IMEI Info")'
    )
    imei_summary = imei_summary[['IMEI','Count','Device Info','First_Use_Date','Last_Use_Date','First_Use_Address','Last_Use_Address']]
    imei_summary['Count'] = imei_summary['Count'].astype(int)
    imei_summary = imei_summary.sort_values(by='Count', ascending=False)

    # ===== معالجة المواقع =====
    site_df = current_df[['Site_Address','Latitude','Longitude','Call_Start_Date']].copy()
    site_group = site_df.groupby('Site_Address').agg(
        Count=('Site_Address','count'),
        Map=('Latitude', lambda x: f'=HYPERLINK("https://www.google.com/maps/search/?api=1&query={x.iloc[0]},{site_df.loc[x.index[0],"Longitude"]}","Map")'),
        First_Use_Date=('Call_Start_Date','min'),
        Last_Use_Date=('Call_Start_Date','max')
    ).reset_index()
    site_group['Count'] = site_group['Count'].astype(int)
    site_group = site_group.sort_values(by='Count', ascending=False)
    site_group = site_group[['Site_Address','Count','Map','First_Use_Date','Last_Use_Date']]

    # ===== تصدير الملف النهائي =====
    output_file = os.path.join(os.path.dirname(current_file), "etisalat_report.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name="calls", index=False)
        imei_summary.to_excel(writer, sheet_name="imei", index=False)
        site_group.to_excel(writer, sheet_name="site", index=False)
        current_df.to_excel(writer, sheet_name="cheet", index=False)

    # ===== تظليل الصف الأول بالأصفر =====
    wb = load_workbook(output_file)
    ws_calls = wb["calls"]
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws_calls[2]:
        cell.fill = yellow_fill

    # ===== تنسيق الأوراق (نفترض format_sheet موجودة) =====
    format_sheet(wb["calls"], header_color="228B22", hyperlink_col=8)
    format_sheet(wb["imei"], header_color="228B22", hyperlink_col=3)
    format_sheet(wb["site"], header_color="228B22", hyperlink_col=3)

    wb.save(output_file)
    messagebox.showinfo("نجاح", f"تم إنشاء تقرير اتصالات\nالملف:\n{output_file}")


def generate_vodafone_report():
    global current_df, current_file

    if current_df is None:
        messagebox.showinfo("مطلوب", "افتح ملف Excel أولاً")
        return

    # ===== الأعمدة المطلوبة في الملف الجديد =====
    required_cols = [
        'B_NUMBER',                 # الرقم الأساسي
        'B_NUMBER_NATIONAL_ID',     # الرقم القومي
        'B_NUMBER_SITE_ADDRESS',    # عنوان موقع الرقم
        'IMEI',                     # IMEI
        'HANDSET_MANUFACTURER',     # مصنع الجهاز
        'HANDSET_MARKETING_NAME',   # اسم الجهاز التسويقي
        'FULL_DATE',                # تاريخ المكالمة
        'SITE_ADDRESS',             # عنوان الموقع الفعلي
        'LATITUDE', 'LONGITUDE',    # إحداثيات
        'SERVICE'                   # نوع الخدمة
    ]

    for col in required_cols:
        if col not in current_df.columns:
            messagebox.showerror("خطأ", f"العمود {col} غير موجود في الملف")
            return

    df = current_df.copy()

    # ===== صفحة CALLS =====
    numbers = df['B_NUMBER'].astype(str)
    freq = numbers.value_counts().reset_index()
    freq.columns = ['B Number','Count']

    sms_count = df[df['SERVICE'].astype(str).str.strip().isin(["Short message MO/PP","Short message MT/PP"])].groupby('B_NUMBER').size().reset_index(name='SMS')

    df_final = freq.merge(
        df[['B_NUMBER','B_NUMBER_NATIONAL_ID','B_NUMBER_SITE_ADDRESS']].drop_duplicates(subset='B_NUMBER'),
        left_on='B Number', right_on='B_NUMBER', how='left'
    )
    df_final = df_final.merge(sms_count, left_on='B Number', right_on='B_NUMBER', how='left')
    df_final['SMS'] = df_final['SMS'].fillna(0).astype(int)
    df_final['B Number Id'] = df_final['B_NUMBER_NATIONAL_ID'].apply(str)

    # إضافة أعمدة First_Call و Last_Call
    df['FULL_DATE'] = pd.to_datetime(df['FULL_DATE'])
    call_dates = df.groupby('B_NUMBER')['FULL_DATE'].agg(First_Call='min', Last_Call='max').reset_index()
    df_final = df_final.merge(call_dates, left_on='B Number', right_on='B_NUMBER', how='left')
    df_final = df_final.drop(columns=['B_NUMBER'])

    # ترتيب الأعمدة النهائية
    df_final = df_final[['B Number','Count','B Number Id','B_NUMBER_SITE_ADDRESS','SMS','First_Call','Last_Call']]
    df_final['B Number'] = df_final['B Number'].astype(str)
    df_final['Count'] = df_final['Count'].astype(int)
    df_final = df_final.sort_values(by='Count', ascending=False)

    # ===== صفحة IMEI =====
    df['IMEI'] = df['IMEI'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    imei_group = df.groupby('IMEI').agg(
        Count=('IMEI','count'),
        Device_Info=('IMEI', lambda x: f'=HYPERLINK("https://www.imei.info/calc/?imei={x.iloc[0]}","IMEI Info")'),
        HANDSET_MANUFACTURER=('HANDSET_MANUFACTURER','first'),
        HANDSET_MARKETING_NAME=('HANDSET_MARKETING_NAME','first'),
        First_Use_Date=('FULL_DATE','min'),
        Last_Use_Date=('FULL_DATE','max')
    ).reset_index()

    first_last_addr = []
    for imei in imei_group['IMEI']:
        sub = df[df['IMEI']==imei].sort_values('FULL_DATE')
        first_addr = sub.iloc[0]['SITE_ADDRESS']
        last_addr = sub.iloc[-1]['SITE_ADDRESS']
        first_last_addr.append((first_addr,last_addr))
    imei_group['First_Use_Address'] = [x[0] for x in first_last_addr]
    imei_group['Last_Use_Address'] = [x[1] for x in first_last_addr]

    imei_group = imei_group[['IMEI','Count','Device_Info','HANDSET_MANUFACTURER','HANDSET_MARKETING_NAME',
                             'First_Use_Date','Last_Use_Date','First_Use_Address','Last_Use_Address']]
    imei_group['Count'] = imei_group['Count'].astype(int)
    imei_group = imei_group.sort_values(by='Count', ascending=False)

    # ===== صفحة SITE =====
    site_df = df[['SITE_ADDRESS','LATITUDE','LONGITUDE','FULL_DATE']].copy()
    site_group = site_df.groupby('SITE_ADDRESS').agg(
        Count=('SITE_ADDRESS','count'),
        Map=('LATITUDE', lambda x: f'=HYPERLINK("https://www.google.com/maps/search/?api=1&query={x.iloc[0]},{site_df.loc[x.index[0],"LONGITUDE"]}","Map")'),
        First_Use_Date=('FULL_DATE','min'),
        Last_Use_Date=('FULL_DATE','max')
    ).reset_index()
    site_group['Count'] = site_group['Count'].astype(int)
    site_group = site_group.sort_values(by='Count', ascending=False)
    site_group = site_group[['SITE_ADDRESS','Count','Map','First_Use_Date','Last_Use_Date']]

    # ===== حفظ التقرير =====
    output_file = os.path.join(os.path.dirname(current_file), "vodafone_report.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_final.to_excel(writer, sheet_name="calls", index=False)
        imei_group.to_excel(writer, sheet_name="imei", index=False)
        site_group.to_excel(writer, sheet_name="site", index=False)
        current_df.to_excel(writer, sheet_name="cheet", index=False)

    # ===== تنسيق الشيتات =====
    wb = load_workbook(output_file)
    format_sheet(wb["calls"], header_color="FF0000")
    format_sheet(wb["imei"], header_color="FF0000", hyperlink_col=3)
    format_sheet(wb["site"], header_color="FF0000", hyperlink_col=3)
    wb.save(output_file)

    messagebox.showinfo("نجاح", f"تم إنشاء تقرير فودافون\nالملف:\n{output_file}")


# ================== أورانج ==================
def generate_orange_report():
    global current_df, current_file
    if current_df is None or current_file is None:
        messagebox.showinfo("مطلوب", "افتح ملف Excel أولاً")
        return

    try:
        df = pd.read_excel(current_file, engine="openpyxl", header=4)
    except Exception as e:
        messagebox.showerror("خطأ في قراءة الملف", str(e))
        return

    required_cols = [
        'TARGET_MSISDN','TARGET_IMEI','TARGET_IMSI','TARGET_IMEI_TYPE','EVENT_START_TIME',
        'CALL_DURATION','EVENT_DIRECTION','OTHER_MSISDN','OTHER_NAME','OTHER_ID',
        'OTHER_ID_TYPE','OTHER_ADDRESS','CELL_ADDRESS','CELL_LAT','CELL_LONG'
    ]
    for col in required_cols:
        if col not in df.columns:
            messagebox.showerror("خطأ", f"العمود {col} غير موجود في الملف")
            return

    numbers = df['OTHER_MSISDN'].astype(str)
    freq = numbers.value_counts().reset_index()
    freq.columns = ['B Number','Count']

    calls_df = freq.merge(
        df[['OTHER_MSISDN','OTHER_NAME','OTHER_ADDRESS','OTHER_ID']].drop_duplicates(subset='OTHER_MSISDN'),
        left_on='B Number', right_on='OTHER_MSISDN', how='left'
    )

    sms_count = df[df['EVENT_DIRECTION'].astype(str).str.strip()=="SMSMT"].groupby('OTHER_MSISDN').size().reset_index(name='SMS')
    calls_df = calls_df.merge(sms_count, left_on='B Number', right_on='OTHER_MSISDN', how='left')
    calls_df['SMS'] = calls_df['SMS'].fillna(0).astype(int)

    # ===== إضافة أعمدة First Call و Last Call =====
    df['EVENT_START_TIME'] = pd.to_datetime(df['EVENT_START_TIME'])
    call_dates = df.groupby('OTHER_MSISDN')['EVENT_START_TIME'].agg(First_Call='min', Last_Call='max').reset_index()
    calls_df = calls_df.merge(call_dates, left_on='B Number', right_on='OTHER_MSISDN', how='left')
    calls_df = calls_df.drop(columns=['OTHER_MSISDN'])

    # ترتيب الأعمدة مع الأعمدة الجديدة في الآخر
    calls_df = calls_df[['B Number','Count','OTHER_NAME','OTHER_ADDRESS','OTHER_ID','SMS','First_Call','Last_Call']]
    calls_df.columns = ['B Number','Count','B Full Name','B Address','B Number id','SMS','First_Call','Last_Call']
    calls_df['B Number'] = calls_df['B Number'].apply(str)
    calls_df['B Number id'] = calls_df['B Number id'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    calls_df['Count'] = calls_df['Count'].astype(int)
    calls_df = calls_df.sort_values(by='Count', ascending=False)

    # ===== صفحة IMEI =====
    df['TARGET_IMEI'] = df['TARGET_IMEI'].apply(lambda x: str(int(x)) if pd.notna(x) else '')
    imei_group = df.groupby('TARGET_IMEI').agg(
        Count=('TARGET_IMEI','count'),
        TARGET_IMEI_TYPE=('TARGET_IMEI_TYPE','first'),
        First_Use_Date=('EVENT_START_TIME','min'),
        Last_Use_Date=('EVENT_START_TIME','max'),
        First_Use_Address=('CELL_ADDRESS','first'),
        Last_Use_Address=('CELL_ADDRESS','last')
    ).reset_index()
    imei_group['Device Info'] = imei_group['TARGET_IMEI'].apply(lambda x: f'=HYPERLINK("https://www.imei.info/calc/?imei={x}","IMEI Info")')
    imei_group = imei_group[['TARGET_IMEI','Count','TARGET_IMEI_TYPE','Device Info','First_Use_Date','Last_Use_Date','First_Use_Address','Last_Use_Address']]
    imei_group.columns = ['IMEI','Count','TARGET_IMEI_TYPE','Device Info','First_Use_Date','Last_Use_Date','First_Use_Address','Last_Use_Address']
    imei_group['Count'] = imei_group['Count'].astype(int)
    imei_group = imei_group.sort_values(by='Count', ascending=False)

    # ===== صفحة SITE =====
    site_df = df.groupby('CELL_ADDRESS').agg(
        Count=('CELL_ADDRESS','count'),
        First_Use_Date=('EVENT_START_TIME','min'),
        Last_Use_Date=('EVENT_START_TIME','max'),
        LAT=('CELL_LAT','first'),
        LON=('CELL_LONG','first')
    ).reset_index()
    site_df['Map'] = site_df.apply(lambda row: f'=HYPERLINK("https://www.google.com/maps/search/?api=1&query={row["LAT"]},{row["LON"]}","Map")' 
                                   if pd.notna(row["LAT"]) and pd.notna(row["LON"]) else '', axis=1)
    site_df = site_df[['CELL_ADDRESS','Count','Map','First_Use_Date','Last_Use_Date']]
    site_df = site_df.sort_values(by='Count', ascending=False)

    # ===== حفظ التقرير =====
    output_file = os.path.join(os.path.dirname(current_file), "orange_report.xlsx")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        calls_df.to_excel(writer, sheet_name="calls", index=False)
        imei_group.to_excel(writer, sheet_name="imei", index=False)
        site_df.to_excel(writer, sheet_name="site", index=False)
        current_df.to_excel(writer, sheet_name="cheet", index=False)

    wb = load_workbook(output_file)
    format_sheet(wb["calls"], header_color="FF6600")
    format_sheet(wb["imei"], header_color="FF6600", hyperlink_col=4)
    format_sheet(wb["site"], header_color="FF6600", hyperlink_col=3)
    wb.save(output_file)
    messagebox.showinfo("نجاح", f"تم إنشاء تقرير أورانج\nالملف:\n{output_file}")

# ================= واجهة المستخدم =================
root = tk.Tk()
root.title("📊 Excel Analyzer Tool")
root.geometry("1300x650")
root.configure(bg="#f9f9f9")

# ===== Header =====
header = tk.Label(root, text="📊 Excel Analyzer Tool", font=("Arial", 22, "bold"), bg="#333", fg="white", pady=10)
header.pack(fill="x")

# ===== Toolbar Buttons =====
btn_frame = tk.Frame(root, bg="#f9f9f9", pady=10)
btn_frame.pack(fill="x")

open_btn = tk.Button(btn_frame, text="📂 فتح ملف Excel", bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), padx=10, pady=5, command=open_excel)
open_btn.pack(side="left", padx=5)

etisalat_btn = tk.Button(btn_frame, text="اتصالات", bg="#228B22", fg="white", font=("Arial", 12, "bold"), padx=10, pady=5, command=generate_etisalat_report)
etisalat_btn.pack(side="left", padx=5)

vodafone_btn = tk.Button(btn_frame, text="فودافون", bg="#FF0000", fg="white", font=("Arial", 12, "bold"), padx=10, pady=5, command=generate_vodafone_report)
vodafone_btn.pack(side="left", padx=5)

orange_btn = tk.Button(btn_frame, text="اورانج", bg="#FF6600", fg="white", font=("Arial", 12, "bold"), padx=10, pady=5, command=generate_orange_report)
orange_btn.pack(side="left", padx=5)

# ===== Treeview =====
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview",
                background="white",
                foreground="black",
                rowheight=28,
                fieldbackground="white",
                font=("Arial", 11))
style.configure("Treeview.Heading",
                font=("Arial", 12, "bold"),
                background="#333",
                foreground="white")
tree = ttk.Treeview(root)
tree.pack(fill="both", expand=True, padx=10, pady=6)

# ===== Status Bar =====
status = tk.Label(root, text="لا يوجد ملف مفتوح", bd=1, relief="sunken", anchor="w", font=("Arial", 10))
status.pack(side="bottom", fill="x")

root.mainloop()
